# fuel/utils/export_utils.py
import io
import csv
from datetime import datetime
from typing import List, Dict, Any

# Conditional imports for heavy dependencies
try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    import qrcode
    from PIL import Image
    QR_AVAILABLE = True
except ImportError:
    QR_AVAILABLE = False

from django.http import HttpResponse
from django.template.loader import render_to_string


class ExportManager:
    """Utility class for exporting data in various formats"""
    
    @staticmethod
    def export_to_csv(data: List[Dict[str, Any]], filename: str = None) -> HttpResponse:
        """Export data to CSV format"""
        if not filename:
            filename = f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        if not data:
            return response
        
        writer = csv.DictWriter(response, fieldnames=data[0].keys())
        writer.writeheader()
        writer.writerows(data)
        
        return response
    
    @staticmethod
    def export_to_excel(data: List[Dict[str, Any]], filename: str = None, sheet_name: str = "Data") -> HttpResponse:
        """Export data to Excel format with styling"""
        if not OPENPYXL_AVAILABLE:
            # Fallback to CSV if openpyxl is not available
            return ExportManager.export_to_csv(data, filename.replace('.xlsx', '.csv') if filename else None)
        
        if not filename:
            filename = f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        if not data:
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            wb.save(response)
            return response
        
        # Add headers
        headers = list(data[0].keys())
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header.replace('_', ' ').title())
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Add data rows
        for row, item in enumerate(data, 2):
            for col, header in enumerate(headers, 1):
                value = item.get(header, '')
                # Format dates
                if isinstance(value, datetime):
                    value = value.strftime('%Y-%m-%d %H:%M:%S')
                ws.cell(row=row, column=col, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
        
        # Save to response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        
        return response
    
    @staticmethod
    def export_to_pdf(data: List[Dict[str, Any]], filename: str = None, title: str = "Export Report") -> HttpResponse:
        """Export data to PDF format"""
        if not REPORTLAB_AVAILABLE:
            # Fallback to CSV if reportlab is not available
            return ExportManager.export_to_csv(data, filename.replace('.pdf', '.csv') if filename else None)
        
        if not filename:
            filename = f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Create PDF document
        doc = SimpleDocTemplate(response, pagesize=A4)
        elements = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1  # Center alignment
        )
        
        # Add title
        elements.append(Paragraph(title, title_style))
        elements.append(Spacer(1, 12))
        
        if not data:
            elements.append(Paragraph("No data available", styles['Normal']))
        else:
            # Prepare table data
            headers = list(data[0].keys())
            table_data = [
                [header.replace('_', ' ').title() for header in headers]
            ]
            
            for item in data:
                row = []
                for header in headers:
                    value = item.get(header, '')
                    if isinstance(value, datetime):
                        value = value.strftime('%Y-%m-%d %H:%M')
                    row.append(str(value))
                table_data.append(row)
            
            # Create table
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            elements.append(table)
        
        # Add footer with timestamp
        elements.append(Spacer(1, 20))
        elements.append(Paragraph(
            f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            styles['Normal']
        ))
        
        doc.build(elements)
        return response


class CouponPrintManager:
    """Utility class for printing fuel coupons"""
    
    @staticmethod
    def generate_coupon_pdf(coupon_data: Dict[str, Any]) -> HttpResponse:
        """Generate a printable fuel coupon PDF"""
        if not REPORTLAB_AVAILABLE or not QR_AVAILABLE:
            # Return error response if required packages are not available
            response = HttpResponse("PDF generation not available. Please install reportlab and qrcode packages.", 
                                  content_type='text/plain', status=503)
            return response
        
        filename = f"fuel_coupon_{coupon_data.get('coupon_number', 'unknown')}.pdf"
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        doc = SimpleDocTemplate(response, pagesize=(4*inch, 6*inch))  # Coupon size
        elements = []
        
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CouponTitle',
            parent=styles['Heading1'],
            fontSize=16,
            alignment=1,
            spaceAfter=10
        )
        elements.append(Paragraph("PARLIAMENT OF ZIMBABWE", title_style))
        elements.append(Paragraph("FUEL COUPON", title_style))
        elements.append(Spacer(1, 10))
        
        # QR Code
        qr = qrcode.QRCode(version=1, box_size=3, border=1)
        qr.add_data(f"COUPON:{coupon_data.get('coupon_number', '')}")
        qr.make(fit=True)
        
        # Coupon details
        details = [
            f"Coupon No: {coupon_data.get('coupon_number', 'N/A')}",
            f"Employee: {coupon_data.get('employee_name', 'N/A')}",
            f"Fuel Type: {coupon_data.get('fuel_type', 'N/A')}",
            f"Liters: {coupon_data.get('liters', 'N/A')}",
            f"Valid Until: {coupon_data.get('valid_until', 'N/A')}",
            f"Issued: {datetime.now().strftime('%Y-%m-%d')}"
        ]
        
        for detail in details:
            elements.append(Paragraph(detail, styles['Normal']))
        
        elements.append(Spacer(1, 10))
        elements.append(Paragraph("Authorized Signature: ________________", styles['Normal']))
        
        doc.build(elements)
        return response
    
    @staticmethod
    def generate_handover_report(handover_data: Dict[str, Any]) -> HttpResponse:
        """Generate handover report PDF"""
        if not REPORTLAB_AVAILABLE:
            # Return error response if reportlab is not available
            response = HttpResponse("PDF generation not available. Please install reportlab package.", 
                                  content_type='text/plain', status=503)
            return response
        
        filename = f"handover_report_{handover_data.get('handover_id', 'unknown')}.pdf"
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        doc = SimpleDocTemplate(response, pagesize=A4)
        elements = []
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'ReportTitle',
            parent=styles['Heading1'],
            fontSize=18,
            alignment=1,
            spaceAfter=20
        )
        
        # Header
        elements.append(Paragraph("PARLIAMENT OF ZIMBABWE", title_style))
        elements.append(Paragraph("FUEL COUPON HANDOVER REPORT", title_style))
        elements.append(Spacer(1, 20))
        
        # Report details
        report_data = [
            ["Handover ID:", handover_data.get('handover_id', 'N/A')],
            ["Date:", handover_data.get('handover_date', 'N/A')],
            ["From:", handover_data.get('from_officer', 'N/A')],
            ["To:", handover_data.get('to_officer', 'N/A')],
            ["Total Coupons:", handover_data.get('total_coupons', 'N/A')],
            ["Total Value:", f"${handover_data.get('total_value', 0):.2f}"]
        ]
        
        table = Table(report_data, colWidths=[2*inch, 4*inch])
        table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 12),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        
        elements.append(table)
        elements.append(Spacer(1, 30))
        
        # Signatures
        sig_data = [
            ["Handed Over By:", "Received By:"],
            ["", ""],
            ["Signature: ________________", "Signature: ________________"],
            ["", ""],
            [f"Date: {datetime.now().strftime('%Y-%m-%d')}", f"Date: {datetime.now().strftime('%Y-%m-%d')}"]
        ]
        
        sig_table = Table(sig_data, colWidths=[3*inch, 3*inch])
        sig_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        
        elements.append(sig_table)
        
        doc.build(elements)
        return response
